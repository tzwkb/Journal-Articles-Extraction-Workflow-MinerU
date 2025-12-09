"""
ä¸»æµç¨‹è„šæœ¬ - ä¿®å¤ç‰ˆ
ä¿®å¤ï¼š
1. å›¾ç‰‡è·¯å¾„å¤„ç†
2. MinerUè¾“å‡ºä½ç½®
3. å›¾ç‰‡å¤åˆ¶é€»è¾‘
"""

import yaml
import sys
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed
from jinja2 import Template
import shutil

from mineru_client import MinerUClient, FileTask
from mineru_parser import MinerUParser
from article_translator import ArticleTranslator
from logger import Logger
from format_converter import FormatConverter
from outline_generator import OutlineGenerator
from path_manager import PathManager

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    from tqdm import tqdm
except ImportError:
    tqdm = None


class DocumentProcessor:
    """æ–‡æ¡£å¤„ç†ä¸»ç±»"""

    def __init__(self, config_path="config.yaml"):
        """
        åˆå§‹åŒ–æ–‡æ¡£å¤„ç†å™¨

        Args:
            config_path: é…ç½®æ–‡ä»¶è·¯å¾„
        """
        # åŠ è½½é…ç½®
        with open(config_path, 'r', encoding='utf-8') as f:
            self.config = yaml.safe_load(f)

        self.logger = Logger()
        self.output_base = Path(self.config['paths']['output_base'])

        # åˆå§‹åŒ–MinerUå®¢æˆ·ç«¯
        self.mineru = MinerUClient(
            api_token=self.config['api']['mineru_token'],
            verify_ssl=False,
            max_retries=5
        )

        # åˆå§‹åŒ–è§£æå™¨ï¼ˆä¿®æ”¹è¾“å‡ºç›®å½•åˆ°output/MinerUï¼‰
        mineru_output_dir = self.output_base / self.config['output']['mineru_folder']
        self.parser = MinerUParser(output_dir=str(mineru_output_dir))

        # åˆå§‹åŒ–æ ¼å¼è½¬æ¢å™¨
        self.converter = FormatConverter(self.config, self.logger, self.output_base)

        # åˆå§‹åŒ–å¤§çº²ç”Ÿæˆå™¨
        self.outline_gen = OutlineGenerator(self.config, self.logger, self.output_base)

        # åˆå§‹åŒ–è·¯å¾„ç®¡ç†å™¨
        self.path_mgr = PathManager(self.config, self.logger)

        # åˆå§‹åŒ–æ–‡ä»¶å¤¹ç»“æ„
        self._init_directories()

    def _init_directories(self):
        """åˆå§‹åŒ–æ‰€éœ€çš„æ–‡ä»¶å¤¹ç»“æ„"""
        input_base = Path(self.config['paths']['input_base'])
        output_base = Path(self.config['paths']['output_base'])
        terminology_folder = Path(self.config['paths']['terminology_folder'])

        # è¾“å‡ºæ–‡ä»¶å¤¹åç§°
        mineru_folder = self.config['output']['mineru_folder']
        html_folder = self.config['output']['html_folder']
        pdf_folder = self.config['output']['pdf_folder']
        docx_folder = self.config['output']['docx_folder']
        cache_folder = self.config['output']['cache_folder']

        # åˆ›å»ºæ‰€æœ‰å¿…è¦çš„ç›®å½•
        folders = [
            input_base,
            terminology_folder,
            output_base / mineru_folder,
            output_base / html_folder,
            output_base / pdf_folder,
            output_base / docx_folder,
            output_base / cache_folder / 'outlines',
        ]

        for folder in folders:
            folder.mkdir(parents=True, exist_ok=True)

        self.logger.info(f"æ–‡ä»¶å¤¹ç»“æ„åˆå§‹åŒ–å®Œæˆ")

    def load_terminology_from_excel(self) -> dict:
        """
        ä» terminology æ–‡ä»¶å¤¹ä¸‹çš„ Excel æ–‡ä»¶åŠ è½½æœ¯è¯­åº“

        Returns:
            æœ¯è¯­å­—å…¸ {"English": "ä¸­æ–‡"}
        """
        terminology_folder = Path(self.config['paths']['terminology_folder'])

        if not terminology_folder.exists():
            self.logger.warning(f"æœ¯è¯­åº“æ–‡ä»¶å¤¹ä¸å­˜åœ¨: {terminology_folder}")
            return {}

        if not load_workbook:
            self.logger.warning("openpyxl æœªå®‰è£…ï¼Œæ— æ³•è¯»å– Excel æœ¯è¯­åº“")
            return {}

        glossary = {}
        excel_files = list(terminology_folder.glob("*.xlsx")) + list(terminology_folder.glob("*.xls"))

        if not excel_files:
            self.logger.warning(f"æœ¯è¯­åº“æ–‡ä»¶å¤¹ä¸­æ²¡æœ‰ Excel æ–‡ä»¶: {terminology_folder}")
            return {}

        self.logger.info(f"æ­£åœ¨åŠ è½½æœ¯è¯­åº“ï¼Œå…± {len(excel_files)} ä¸ª Excel æ–‡ä»¶...")

        for excel_file in excel_files:
            try:
                workbook = load_workbook(excel_file, read_only=True, data_only=True)

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]

                    if sheet.max_row <= 1:
                        continue

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 2 and row[0] and row[1]:
                            english_term = str(row[0]).strip()
                            chinese_term = str(row[1]).strip()

                            if english_term and chinese_term:
                                glossary[english_term] = chinese_term

                workbook.close()
                self.logger.info(f"  å·²åŠ è½½: {excel_file.name} - {len(glossary)} ä¸ªæœ¯è¯­")

            except Exception as e:
                self.logger.error(f"åŠ è½½ Excel æ–‡ä»¶å¤±è´¥: {excel_file.name} - {str(e)}")

        self.logger.success(f"æœ¯è¯­åº“åŠ è½½å®Œæˆï¼Œå…± {len(glossary)} ä¸ªæœ¯è¯­")
        return glossary

    def batch_process(self):
        """
        æ‰¹é‡å¤„ç† input æ–‡ä»¶å¤¹ä¸­çš„æ‰€æœ‰ PDF æ–‡ä»¶ï¼ˆå¤šæ–‡ä»¶å¹¶å‘ï¼‰
        """
        self.logger.info("=" * 60)
        self.logger.info("æ‰¹é‡å¤„ç†æ¨¡å¼")
        self.logger.info("=" * 60)

        # 1. æ‰«æè¾“å…¥æ–‡ä»¶
        file_list = self.path_mgr.scan_input_files()

        if not file_list:
            self.logger.error("æ²¡æœ‰æ‰¾åˆ°è¦å¤„ç†çš„ PDF æ–‡ä»¶")
            return

        # 2. åŠ è½½å…¨å±€æœ¯è¯­åº“ï¼ˆä» Excelï¼‰
        excel_glossary = self.load_terminology_from_excel()

        # 3. å¤šæ–‡ä»¶å¹¶å‘å¤„ç†
        max_workers = self.config['concurrency']['max_files']
        self.logger.info(f"å¼€å§‹å¹¶å‘å¤„ç†ï¼Œå¹¶å‘æ•°: {max_workers}")

        success_count = 0
        failure_count = 0
        results = []

        with ProcessPoolExecutor(max_workers=max_workers) as executor:
            future_to_file = {
                executor.submit(self._process_single_file, relative_path, pdf_path, excel_glossary):
                (relative_path, pdf_path)
                for relative_path, pdf_path in file_list
            }

            if tqdm:
                future_iterator = tqdm(as_completed(future_to_file), total=len(file_list), desc="å¤„ç†è¿›åº¦")
            else:
                future_iterator = as_completed(future_to_file)

            for future in future_iterator:
                relative_path, pdf_path = future_to_file[future]
                try:
                    result = future.result()
                    if result['success']:
                        success_count += 1
                        self.logger.success(f"âœ“ å®Œæˆ: {relative_path}")
                    else:
                        failure_count += 1
                        self.logger.error(f"âœ— å¤±è´¥: {relative_path} - {result.get('error', 'Unknown error')}")
                    results.append(result)
                except Exception as e:
                    failure_count += 1
                    self.logger.error(f"âœ— å¤±è´¥: {relative_path} - {str(e)}")
                    results.append({'success': False, 'file': relative_path, 'error': str(e)})

        # 4. è¾“å‡ºæ±‡æ€»
        self.logger.info("=" * 60)
        self.logger.info(f"æ‰¹é‡å¤„ç†å®Œæˆï¼")
        self.logger.info(f"  æˆåŠŸ: {success_count} ä¸ªæ–‡ä»¶")
        self.logger.info(f"  å¤±è´¥: {failure_count} ä¸ªæ–‡ä»¶")
        self.logger.info("=" * 60)

        return results

    def _process_single_file(self, relative_path: str, pdf_path: str, excel_glossary: dict) -> dict:
        """
        å¤„ç†å•ä¸ª PDF æ–‡ä»¶ï¼ˆç”¨äºå¤šè¿›ç¨‹è°ƒç”¨ï¼‰

        Args:
            relative_path: ç›¸å¯¹è·¯å¾„
            pdf_path: PDF ç»å¯¹è·¯å¾„
            excel_glossary: Excel æœ¯è¯­åº“

        Returns:
            å¤„ç†ç»“æœå­—å…¸
        """
        try:
            output_paths = self.path_mgr.get_output_paths(relative_path)
            self.run(pdf_path, output_paths, excel_glossary)

            return {
                'success': True,
                'file': relative_path,
                'output_paths': {k: str(v) for k, v in output_paths.items()}
            }
        except Exception as e:
            return {
                'success': False,
                'file': relative_path,
                'error': str(e)
            }

    def run(self, pdf_path: str, output_paths: dict = None, excel_glossary: dict = None):
        """
        è¿è¡Œå®Œæ•´æµç¨‹

        Args:
            pdf_path: PDFæ–‡ä»¶è·¯å¾„
            output_paths: è‡ªå®šä¹‰è¾“å‡ºè·¯å¾„å­—å…¸ï¼ˆå¯é€‰ï¼‰
            excel_glossary: Excelæœ¯è¯­åº“ï¼ˆå¯é€‰ï¼‰
        """
        self.logger.info("=" * 60)
        self.logger.info("å¼€å§‹å¤„ç†æ–‡æ¡£")
        self.logger.info("=" * 60)

        try:
            # æ­¥éª¤1: ç”Ÿæˆå¤§çº²
            outline = self.outline_gen.generate_outline(pdf_path, output_paths)

            # æ­¥éª¤2: MinerUè§£æ
            content_list, extract_dir = self.parse_with_mineru(pdf_path, output_paths)

            # æ­¥éª¤3: ä½¿ç”¨ Excel æœ¯è¯­åº“ï¼ˆä¸ä½¿ç”¨ AI ç”Ÿæˆçš„æœ¯è¯­ï¼‰
            combined_glossary = excel_glossary or {}
            
            if combined_glossary:
                self.logger.info(f"æœ¯è¯­åº“åŠ è½½å®Œæˆ: {len(combined_glossary)} ä¸ªæœ¯è¯­")
            else:
                self.logger.warning("æœªæ‰¾åˆ°æœ¯è¯­åº“ï¼Œå°†ä¸è¿›è¡Œæœ¯è¯­é¢„æ›¿æ¢")

            # æ­¥éª¤4: åˆå§‹åŒ–ç¿»è¯‘å™¨
            translator = ArticleTranslator(
                api_key=self.config['api']['translation_api_key'],
                api_url=self.config['api']['translation_api_base_url'],
                model=self.config['api']['translation_api_model'],
                glossary=combined_glossary,
                case_sensitive=False,
                whole_word_only=True,
                config=self.config
            )

            # æ­¥éª¤5: å¤„ç†å†…å®¹å¹¶ç¿»è¯‘
            original_html, translated_html = self.process_content(
                content_list, outline, translator, extract_dir, output_paths
            )

            # æ­¥éª¤6: å¯¼å‡ºæ ¼å¼
            self.converter.export_formats(original_html, translated_html, output_paths)

            self.logger.info("=" * 60)
            self.logger.success("å¤„ç†å®Œæˆï¼")
            self.logger.info("=" * 60)

        except Exception as e:
            self.logger.error(f"å¤„ç†å¤±è´¥: {str(e)}")
            import traceback
            traceback.print_exc()
            raise

    def parse_with_mineru(self, pdf_path: str, output_paths: dict = None) -> tuple:
        """
        ä½¿ç”¨MinerUè§£æPDF

        Args:
            pdf_path: PDFæ–‡ä»¶è·¯å¾„
            output_paths: è‡ªå®šä¹‰è¾“å‡ºè·¯å¾„å­—å…¸ï¼ˆå¯é€‰ï¼‰

        Returns:
            (content_list, extract_dir) - å†…å®¹åˆ—è¡¨å’Œè§£å‹ç›®å½•
        """
        self.logger.info("\n>>> æ­¥éª¤2: ä½¿ç”¨MinerUè§£æPDF...")

        # ç¡®å®šZIPä¿å­˜è·¯å¾„ï¼ˆoutput/MinerU/ç›¸å¯¹è·¯å¾„ï¼‰
        if output_paths and 'mineru' in output_paths:
            expected_zip = Path(output_paths['mineru'])
        else:
            mineru_folder = self.config['output']['mineru_folder']
            mineru_dir = self.output_base / mineru_folder
            pdf_name = Path(pdf_path).stem
            expected_zip = mineru_dir / f"{pdf_name}_result.zip"

        expected_zip.parent.mkdir(parents=True, exist_ok=True)

        # æ£€æŸ¥æ˜¯å¦å·²æœ‰è§£æç»“æœ
        if expected_zip.exists():
            self.logger.info("å‘ç°å·²æœ‰MinerUè§£æç»“æœï¼Œç›´æ¥åŠ è½½...")
            parsed = self.parser.parse_zip_result(
                str(expected_zip),
                source_file_name=Path(pdf_path).name
            )
            # è·å–è§£å‹ç›®å½•
            extract_dir = self.parser.output_dir / Path(expected_zip).stem
            self.logger.success(f"è§£æç»“æœå·²åŠ è½½: {len(parsed.json_content)} ä¸ªå†…å®¹å—")
            return parsed.json_content, str(extract_dir)

        # ä¸Šä¼ å¹¶è§£æ
        file_task = FileTask(
            file_name=Path(pdf_path).name,
            file_path=pdf_path,
            data_id=Path(pdf_path).stem
        )

        self.logger.info("æ­£åœ¨ä¸Šä¼ PDFåˆ°MinerU...")
        batch_id, _ = self.mineru.batch_upload_files([file_task])

        self.logger.info("ç­‰å¾…MinerUè§£æå®Œæˆ...")
        results = self.mineru.wait_for_completion(batch_id, poll_interval=10)

        # ä¸‹è½½ç»“æœåˆ°æŒ‡å®šä½ç½®
        downloaded = self.mineru.download_all_results(results, str(expected_zip.parent))

        # è·å–ä¸‹è½½çš„zipæ–‡ä»¶è·¯å¾„
        zip_path = list(downloaded.values())[0]

        # å¦‚æœä¸‹è½½ä½ç½®ä¸æ˜¯ç›®æ ‡ä½ç½®ï¼Œç§»åŠ¨æ–‡ä»¶
        if Path(zip_path) != expected_zip:
            shutil.move(zip_path, str(expected_zip))

        # è§£æZIP
        parsed = self.parser.parse_zip_result(
            str(expected_zip),
            source_file_name=Path(pdf_path).name
        )

        # è·å–è§£å‹ç›®å½•
        extract_dir = self.parser.output_dir / Path(expected_zip).stem

        self.logger.success(f"è§£æå®Œæˆ: {len(parsed.json_content)} ä¸ªå†…å®¹å—")
        return parsed.json_content, str(extract_dir)

    def process_content(
        self,
        content_list: list,
        outline: dict,
        translator: ArticleTranslator,
        extract_dir: str,
        output_paths: dict = None
    ) -> tuple:
        """
        å¤„ç†å†…å®¹å¹¶ç¿»è¯‘

        Args:
            content_list: MinerUè¿”å›çš„content_list
            outline: æ–‡æ¡£å¤§çº²
            translator: ç¿»è¯‘å™¨å®ä¾‹
            extract_dir: MinerUè§£å‹ç›®å½•
            output_paths: è¾“å‡ºè·¯å¾„å­—å…¸

        Returns:
            (original_html, translated_html) å…ƒç»„
        """
        self.logger.info("\n>>> æ­¥éª¤3: å¤„ç†å†…å®¹å¹¶ç¿»è¯‘...")

        # æŒ‰é¡µåˆ†ç»„
        pages = {}
        for item in content_list:
            page_idx = item.get('page_idx', 0)
            if page_idx not in pages:
                pages[page_idx] = []
            pages[page_idx].append(item)

        self.logger.info(f"å…± {len(pages)} é¡µ")

        # å¤„ç†å›¾ç‰‡ï¼šå¤åˆ¶åˆ°HTMLç›®å½•å¹¶æ›´æ–°è·¯å¾„
        self._process_images(content_list, extract_dir, output_paths)

        # æ”¶é›†ç¿»è¯‘ä»»åŠ¡
        tasks = []
        for page_idx in sorted(pages.keys()):
            items = pages[page_idx]
            context = self._get_chapter_context(page_idx, outline)

            for item in items:
                if item['type'] in ['header', 'footer', 'page_number']:
                    continue

                if item['type'] == 'text' and item.get('text'):
                    tasks.append((item, 'text_zh', item['text'], context))

                if item['type'] == 'image' and item.get('image_caption'):
                    caption_text = ' '.join(item['image_caption'])
                    tasks.append((item, 'caption_zh', caption_text, context))

        self.logger.info(f"å…±æ”¶é›† {len(tasks)} ä¸ªç¿»è¯‘ä»»åŠ¡ï¼Œå¼€å§‹å¹¶å‘ç¿»è¯‘...")

        # æ‰¹é‡å¹¶å‘ç¿»è¯‘
        translation_tasks = [(text, context) for _, _, text, context in tasks]
        translations = translator.translate_batch(translation_tasks)

        # èµ‹å€¼ç¿»è¯‘ç»“æœ
        for i, (item, field_name, _, _) in enumerate(tasks):
            item[field_name] = translations[i]

            if (i + 1) % max(1, len(tasks) // 10) == 0:
                progress = (i + 1) * 100 // len(tasks)
                self.logger.info(f"  ç¿»è¯‘è¿›åº¦: {i + 1}/{len(tasks)} ({progress}%)")

        self.logger.success(f"ç¿»è¯‘å®Œæˆ: {len(tasks)} ä¸ªå†…å®¹å—")

        # ç”ŸæˆHTML
        self.logger.info("æ­£åœ¨ç”ŸæˆHTML...")
        original_html = self._render_html(pages, language='en')
        translated_html = self._render_html(pages, language='zh')

        self.logger.success("HTMLå·²ç”Ÿæˆ")

        return original_html, translated_html

    def _process_images(self, content_list: list, extract_dir: str, output_paths: dict = None):
        """
        å¤„ç†å›¾ç‰‡ï¼šå¤åˆ¶å›¾ç‰‡åˆ°HTMLè¾“å‡ºç›®å½•å¹¶æ›´æ–°è·¯å¾„

        Args:
            content_list: å†…å®¹åˆ—è¡¨
            extract_dir: MinerUè§£å‹ç›®å½•
            output_paths: è¾“å‡ºè·¯å¾„å­—å…¸
        """
        extract_dir = Path(extract_dir)
        source_images_dir = extract_dir / "images"

        if not source_images_dir.exists():
            self.logger.warning(f"æœªæ‰¾åˆ°å›¾ç‰‡ç›®å½•: {source_images_dir}")
            return

        # ç¡®å®šç›®æ ‡å›¾ç‰‡ç›®å½•ï¼ˆç»Ÿä¸€æ”¾åœ¨ output/HTML/images/ï¼‰
        html_folder = self.config['output']['html_folder']
        html_base_dir = self.output_base / html_folder
        
        if output_paths and 'html_original' in output_paths:
            # ä½¿ç”¨ä¸ HTML æ–‡ä»¶ç›¸åŒçš„ç›®å½•å±‚çº§
            html_dir = Path(output_paths['html_original']).parent
        else:
            html_dir = html_base_dir

        target_images_dir = html_dir / "images"
        target_images_dir.mkdir(parents=True, exist_ok=True)

        self.logger.info(f"æ­£åœ¨å¤åˆ¶å›¾ç‰‡: {source_images_dir} -> {target_images_dir}")

        # å¤åˆ¶å›¾ç‰‡å¹¶æ›´æ–°è·¯å¾„
        copied_count = 0
        for item in content_list:
            if item.get('type') == 'image' and item.get('img_path'):
                img_rel_path = item['img_path']
                source_img = extract_dir / img_rel_path

                if source_img.exists():
                    img_filename = Path(img_rel_path).name
                    target_img = target_images_dir / img_filename

                    # å¤åˆ¶å›¾ç‰‡
                    shutil.copy2(source_img, target_img)

                    # æ›´æ–°è·¯å¾„ï¼š
                    # 1. ç›¸å¯¹è·¯å¾„ç”¨äº HTMLï¼ˆimages/xxx.jpgï¼‰
                    # 2. ç»å¯¹è·¯å¾„ç”¨äº PDF/DOCX è½¬æ¢ï¼ˆå­˜å‚¨åœ¨ img_path_absoluteï¼‰
                    item['img_path'] = f"images/{img_filename}"
                    item['img_path_absolute'] = str(target_img.absolute())
                    copied_count += 1
                else:
                    self.logger.warning(f"å›¾ç‰‡æ–‡ä»¶ä¸å­˜åœ¨: {source_img}")

        if copied_count > 0:
            self.logger.success(f"å·²å¤åˆ¶ {copied_count} å¼ å›¾ç‰‡")
        else:
            self.logger.warning("æœªæ‰¾åˆ°ä»»ä½•å›¾ç‰‡æ–‡ä»¶")

    def _get_chapter_context(self, page_idx: int, outline: dict) -> dict:
        """è·å–é¡µé¢å¯¹åº”çš„ç« èŠ‚ä¸Šä¸‹æ–‡"""
        # ç¡®ä¿ page_idx æ˜¯æ•´æ•°
        try:
            page_num = int(page_idx)
        except (ValueError, TypeError):
            return {}

        for chapter in outline.get('structure', []):
            pages = chapter.get('pages', [])
            if len(pages) >= 2:
                try:
                    # ç¡®ä¿ start å’Œ end ä¹Ÿæ˜¯æ•´æ•°
                    start = int(pages[0])
                    end = int(pages[1])
                    if start <= page_num <= end:
                        return {
                            'chapter_title': chapter.get('title', ''),
                            'chapter_summary': chapter.get('summary', ''),
                            'keywords': chapter.get('keywords', [])
                        }
                except (ValueError, TypeError, IndexError):
                    continue
        return {}

    def _render_html(self, pages: dict, language: str) -> str:
        """æ¸²æŸ“HTML"""
        with open('page_template.html', 'r', encoding='utf-8') as f:
            template = Template(f.read())

        return template.render(pages=pages, language=language)


def main():
    """å‘½ä»¤è¡Œå…¥å£"""
    if len(sys.argv) == 1:
        interactive_mode()
        return

    if sys.argv[1] in ["--batch", "-b", "--interactive", "-i"]:
        interactive_mode()
    else:
        print(f"âŒ æœªçŸ¥å‚æ•°: {sys.argv[1]}")
        print("ä½¿ç”¨ 'python main.py -h' æŸ¥çœ‹å¸®åŠ©")
        sys.exit(1)

def interactive_mode():
    """äº¤äº’å¼å‘½ä»¤è¡Œç•Œé¢"""
    processor = DocumentProcessor()

    while True:
        print("\n" + "="*60)
        print("  MinerU æ–‡æ¡£ç¿»è¯‘å·¥å…· - äº¤äº’æ¨¡å¼")
        print("="*60)
        print("\nè¯·é€‰æ‹©æ“ä½œï¼š")
        print("  [1] æ‰¹é‡å¤„ç†ï¼ˆé€’å½’æ‰«æ input/ æ–‡ä»¶å¤¹ï¼‰")
        print("  [2] æŸ¥çœ‹é…ç½®ä¿¡æ¯")
        print("  [3] æŸ¥çœ‹è¾“å…¥æ–‡ä»¶åˆ—è¡¨")
        print("  [4] æ¸…é™¤ç¼“å­˜")
        print("  [0] é€€å‡º")
        print()

        choice = input("è¯·è¾“å…¥é€‰é¡¹ [0-4]: ").strip()

        if choice == "0":
            print("\nå†è§ï¼")
            break
        elif choice == "1":
            batch_mode_interactive(processor)
        elif choice == "2":
            show_config(processor)
        elif choice == "3":
            show_input_files(processor)
        elif choice == "4":
            clear_cache(processor)
        else:
            print("âŒ æ— æ•ˆé€‰é¡¹ï¼Œè¯·é‡æ–°é€‰æ‹©")


def batch_mode_interactive(processor):
    """æ‰¹é‡å¤„ç†äº¤äº’æ¨¡å¼"""
    print("\n" + "-"*60)
    print("  æ‰¹é‡å¤„ç†æ¨¡å¼")
    print("-"*60)

    file_list = processor.path_mgr.scan_input_files()

    if not file_list:
        print("\nâŒ input/ æ–‡ä»¶å¤¹ä¸­æ²¡æœ‰æ‰¾åˆ° PDF æ–‡ä»¶")
        print("   è¯·å…ˆå°† PDF æ–‡ä»¶æ”¾å…¥ input/ æ–‡ä»¶å¤¹")
        input("\næŒ‰å›è½¦é”®ç»§ç»­...")
        return

    print(f"\næ‰¾åˆ° {len(file_list)} ä¸ª PDF æ–‡ä»¶:")
    for i, (rel_path, abs_path) in enumerate(file_list[:10], 1):
        print(f"  {i}. {rel_path}")

    if len(file_list) > 10:
        print(f"  ... è¿˜æœ‰ {len(file_list) - 10} ä¸ªæ–‡ä»¶")

    print(f"\nå¹¶å‘é…ç½®:")
    print(f"  - æ–‡ä»¶å¹¶å‘æ•°: {processor.config['concurrency']['max_files']}")
    print(f"  - ç¿»è¯‘å¹¶å‘æ•°: {processor.config['concurrency']['initial_translation_workers']} (åˆå§‹)")

    confirm = input(f"\nç¡®è®¤å¼€å§‹æ‰¹é‡å¤„ç†ï¼Ÿ[y/N]: ").strip().lower()

    if confirm != 'y':
        print("å·²å–æ¶ˆ")
        return

    try:
        print("\nå¼€å§‹æ‰¹é‡å¤„ç†...")
        processor.batch_process()
        print("\nâœ“ æ‰¹é‡å¤„ç†å®Œæˆï¼")
    except Exception as e:
        print(f"\nâŒ æ‰¹é‡å¤„ç†å¤±è´¥: {str(e)}")

    input("\næŒ‰å›è½¦é”®ç»§ç»­...")


def show_config(processor):
    """æ˜¾ç¤ºé…ç½®ä¿¡æ¯"""
    print("\n" + "-"*60)
    print("  å½“å‰é…ç½®ä¿¡æ¯")
    print("-"*60)

    config = processor.config

    print("\nğŸ“¡ API é…ç½®:")
    print(f"  MinerU Token: {'å·²é…ç½®' if config['api']['mineru_token'] != 'YOUR_MINERU_TOKEN' else 'âŒ æœªé…ç½®'}")
    print(f"  Translation API: {config['api']['translation_api_base_url']}")
    print(f"  Translation Model: {config['api']['translation_api_model']}")

    print("\nğŸ”„ å¹¶å‘é…ç½®:")
    print(f"  æ–‡ä»¶å¹¶å‘æ•°: {config['concurrency']['max_files']}")
    print(f"  ç¿»è¯‘å¹¶å‘æ•°: {config['concurrency']['initial_translation_workers']}-{config['concurrency']['max_translation_workers']}")

    input("\næŒ‰å›è½¦é”®ç»§ç»­...")


def show_input_files(processor):
    """æ˜¾ç¤ºè¾“å…¥æ–‡ä»¶åˆ—è¡¨"""
    print("\n" + "-"*60)
    print("  è¾“å…¥æ–‡ä»¶åˆ—è¡¨")
    print("-"*60)

    file_list = processor.path_mgr.scan_input_files()

    if not file_list:
        print("\nâŒ input/ æ–‡ä»¶å¤¹ä¸­æ²¡æœ‰æ‰¾åˆ° PDF æ–‡ä»¶")
    else:
        print(f"\næ‰¾åˆ° {len(file_list)} ä¸ª PDF æ–‡ä»¶:\n")
        for i, (rel_path, abs_path) in enumerate(file_list, 1):
            file_size = Path(abs_path).stat().st_size / (1024 * 1024)
            print(f"  {i:3d}. {rel_path:50s} ({file_size:.1f} MB)")

    input("\næŒ‰å›è½¦é”®ç»§ç»­...")


def clear_cache(processor):
    """æ¸…é™¤ç¼“å­˜"""
    print("\n" + "-"*60)
    print("  æ¸…é™¤ç¼“å­˜")
    print("-"*60)

    cache_dir = processor.output_base / "cache"

    if not cache_dir.exists():
        print("\næ²¡æœ‰ç¼“å­˜éœ€è¦æ¸…é™¤")
        input("\næŒ‰å›è½¦é”®ç»§ç»­...")
        return

    total_size = 0
    file_count = 0
    for file in cache_dir.rglob("*"):
        if file.is_file():
            total_size += file.stat().st_size
            file_count += 1

    print(f"\nç¼“å­˜ç»Ÿè®¡:")
    print(f"  æ–‡ä»¶æ•°: {file_count}")
    print(f"  æ€»å¤§å°: {total_size / (1024 * 1024):.1f} MB")

    confirm = input("\nç¡®è®¤æ¸…é™¤æ‰€æœ‰ç¼“å­˜ï¼Ÿ[y/N]: ").strip().lower()

    if confirm != 'y':
        print("å·²å–æ¶ˆ")
        input("\næŒ‰å›è½¦é”®ç»§ç»­...")
        return

    try:
        shutil.rmtree(cache_dir)
        cache_dir.mkdir(parents=True, exist_ok=True)
        print("\nâœ“ ç¼“å­˜å·²æ¸…é™¤")
    except Exception as e:
        print(f"\nâŒ æ¸…é™¤å¤±è´¥: {str(e)}")

    input("\næŒ‰å›è½¦é”®ç»§ç»­...")


if __name__ == "__main__":
    main()